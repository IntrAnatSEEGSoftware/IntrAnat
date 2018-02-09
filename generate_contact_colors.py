#! /usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl, json, os
import pdb
import difflib #to get closest match



class generate_contact_colors:
  """to present results of cognition/stimulation"""
  
  def __init__(self):
      
      self.output = {}
  
  def from_excel_files(self,filename,save=False):
    
    excel_data = openpyxl.load_workbook(filename)
    
    worksheet_names = excel_data.get_sheet_names()
    
    output_info = {worksheet_names[x]:{} for x in range(len(worksheet_names))}
    
    #pour les fréquences
    freq_possible = ['1 Hz', '50 Hz']
    amplitude_possible = []
    
    for worksheet_index in worksheet_names:
        datasheeti = excel_data[worksheet_index]
        
        #always assume first line is "title condition" and first column is "contact or bipole" as written in the IntrAnat Electrodes software
        title_condi={}
        contact_labels = {}
        for i_row in range(1,datasheeti.max_row+1):
            skip_this_line = 0
            for i_column in range(1,datasheeti.max_column+1):
        
        
                #print(i_row,i_column)
                actual_cell = datasheeti.cell(row=i_row, column = i_column)
                backcolor = actual_cell.fill.start_color.index
                fontcolor = actual_cell.font.color.rgb
                if i_row == 1:
                    title_condi.update({actual_cell.value:{}})
                    title_condi[actual_cell.value].update({'title_backcolor':self.hexa2irgbtupple(backcolor),'title_fontcolor':self.hexa2irgbtupple(fontcolor)})
                else:
                    second_label = datasheeti.cell(row=i_row, column = 2)
                    if i_column == 1: 
                       if actual_cell.value not in contact_labels.keys():
                         contact_labels.update({actual_cell.value:{'cell':{},'line':{'backcolor':self.hexa2irgbtupple(backcolor),'fontcolor':self.hexa2irgbtupple(fontcolor)}}}) #second_label.value:{}
                    else:
                       #find row and column title
                       row_title_cell = datasheeti.cell(row = i_row, column = 1)
                       column_title_cell = datasheeti.cell(row = 1, column = i_column)
                       row_title = row_title_cell.value
                       column_title = column_title_cell.value
                       
                       if i_column == 2:
                           if actual_cell.value in contact_labels[row_title]['cell'].keys():
                             if 'Type of response' in contact_labels[row_title]['cell'][actual_cell.value].keys():
                                 #bipole already stimulated with this "frequency/task", data overwrite only if there was a clinical response
                                 #should check the mA ?
                                 if (contact_labels[row_title]['cell'][actual_cell.value]['Type of response']['value'] != u'Absent') and (contact_labels[row_title]['cell'][actual_cell.value]['Type of response'] != 0):
                                   skip_this_line = 1
                                   continue
                                 else:
                                   contact_labels[row_title]['cell'][actual_cell.value].update({})
                             else:
                                 pdb.set_trace()
                           else:
                             contact_labels[row_title]['cell'].update({actual_cell.value:{}}) 
                             #[second_label.value].update({column_title:{'value':actual_cell.value,'backcolor':self.hexa2irgbtupple(backcolor),'fontcolor':self.hexa2irgbtupple(fontcolor)}})  
                       else:
                         if skip_this_line == 0:  
                           contact_labels[row_title]['cell'][second_label.value].update({column_title:{'value':actual_cell.value,'backcolor':self.hexa2irgbtupple(backcolor),'fontcolor':self.hexa2irgbtupple(fontcolor)}})
    
    if save:
        path_to_save = os.path.dirname(filename)
        filename_wh_ext = os.path.basename(filename).split('.')[0]
        new_filename = path_to_save + os.path.sep + filename_wh_ext + '.json'
        fout = open(new_filename,'w')
        fout.write(json.dumps({'title':title_condi,'contacts':contact_labels}))
        fout.close()           
        
    else:
        return (title_condi,contact_labels)
  
  def hexa2irgbtupple(self,string_hexa):
      
      try:
        rgbtupple = (int(string_hexa[0:2],16),int(string_hexa[2:4],16),int(string_hexa[4:6],16),int(string_hexa[6:8],16))
      except:
        rgbtupple = (255,0,0,0) #default is black

      return rgbtupple
  